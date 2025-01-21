import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import plotly.express as px
import os
from datetime import date
from openpyxl import load_workbook
from pyvis.network import Network
import tempfile
import streamlit.components.v1 as components
import json

# ------------------------------------------
#         CONSTANTS / FILE REFERENCES
# ------------------------------------------
FILE_NAME = "VCR - All Enacted Law & Legislative Tracker.xlsx"
SHEET_NAME = "Enacted Federal Law (Ex. J.Res."
DEFAULT_AUTHOR_NAME = "Sullivan"  # Adjust if your data uses a different string

# ==========================================
#         DATA LOADING & PREP
# ==========================================
@st.cache_data
def load_data() -> pd.DataFrame:
    """
    Load, clean, and structure the spreadsheet data.
    """
    if not os.path.exists(FILE_NAME):
        st.error(f"File '{FILE_NAME}' not found in the current directory.")
        return pd.DataFrame()

    # Load data from Excel
    df = pd.read_excel(FILE_NAME, sheet_name=SHEET_NAME, engine="openpyxl")

    # Keep specific columns and rename for readability
    df = df[
        [
            "Author(s)",
            "Original Introduction Date:",
            "Main policy topic",
            "Current Link (Inc. Amndt, if applicable)",
            "Method of Enactment",
        ]
    ].copy()
    df.columns = ["Authors", "Date", "Policy Area", "Title and Link", "Enactment Method"]

    # Convert Date column to datetime
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # Use openpyxl to extract actual hyperlink targets from the relevant column
    workbook = load_workbook(FILE_NAME)
    sheet = workbook[SHEET_NAME]
    links = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
        cell = row[0]
        links.append(cell.hyperlink.target if cell.hyperlink else None)

    df["Link"] = links

    # Extract plain text for "Title" by removing embedded URLs from the string
    df["Title"] = df["Title and Link"].str.replace(r"http[^\s]+", "", regex=True).str.strip()

    # Explode authors by comma to facilitate filtering (e.g., "Sen. A, Sen. B" -> 2 rows)
    df = df.assign(Author=df["Authors"].str.split(",")).explode("Author")
    df["Author"] = df["Author"].str.strip()  # remove extra spaces

    return df

@st.cache_data
def get_filtered_data() -> pd.DataFrame:
    """
    Returns the cleaned data with valid dates only.
    """
    data = load_data()
    if not data.empty:
        # Drop rows with no valid Date
        data = data.dropna(subset=["Date"]).reset_index(drop=True)
    return data

# ==========================================
#         HELPER FUNCTIONS
# ==========================================
def generate_scatter_plot(
    data: pd.DataFrame,
    x_col: str,
    y_col: str,
    color_col: str,
    title: str,
    text_size: int = 12,
    annotate_points: bool = True
) -> px.scatter:
    """
    Generates a scatter plot with:
      - Hover info (including Method of Enactment).
      - Annotations that create clickable links (if annotate_points=True).
      - Consistent styling.
    """
    fig = px.scatter(
        data,
        x=x_col,
        y=y_col,
        color=color_col,
        size=[10] * len(data),  # fixed orb size
        hover_name="Title",
        hover_data={
            "Date": True,
            "Link": True,
            "Enactment Method": True,
            color_col: True,
        },
        labels={
            "Policy Area": "Policy Area",
            "Date": "Date Introduced",
            "Author": "Author",
            "Enactment Method": "Method of Enactment",
        },
        title=title,
    )
    fig.update_layout(
        autosize=True,
        height=700,
        font=dict(size=text_size),
        margin=dict(l=40, r=40, t=80, b=40),
    )

    # Optionally add clickable annotations for each data point
    if annotate_points:
        for i, row in data.iterrows():
            link = row.get("Link", None)
            x_val = row[x_col]
            y_val = row[y_col]
            title_text = row["Title"] or ""

            if pd.notna(link):
                annotation_text = f'<a href="{link}" target="_blank">{title_text}</a>'
            else:
                annotation_text = title_text  # no link if missing

            if annotation_text.strip():
                fig.add_annotation(
                    x=x_val,
                    y=y_val,
                    text=annotation_text,
                    showarrow=False,
                    yshift=10,  # shift label upward
                    font=dict(size=text_size - 2, color="blue"),
                )

    return fig

def display_results_table(df: pd.DataFrame):
    """
    Displays a summary and a nicely formatted table of the filtered results.
    """
    count = len(df)
    st.write(f"**Total matching records:** {count}")
    if count > 0:
        columns_to_show = [
            "Author",
            "Date",
            "Policy Area",
            "Enactment Method",
            "Title",
            "Link",
        ]
        st.dataframe(df[columns_to_show].reset_index(drop=True))
    else:
        st.info("No records match your selection.")

def create_network_graph(data: pd.DataFrame) -> Network:
    """
    Creates and returns an interactive PyVis force-directed Network graph
    where each Author, Policy Area, and Bill Title is a node, with edges
    showing relationships (Author -> Bill, Bill -> Policy Area).

    CHANGE: Added the Bill's Date to the node tooltip.
    """
    net = Network(height="700px", width="100%", bgcolor="#222222", font_color="white")
    net.force_atlas_2based()  # More stable

    added_nodes = set()

    for idx, row in data.iterrows():
        bill_title = row["Title"]
        author = row["Author"]
        policy_area = row["Policy Area"]
        link = row.get("Link", None)

        # We'll convert the date to a short string (YYYY-MM-DD) if available
        bill_date = row.get("Date", None)
        if pd.notna(bill_date):
            date_str = bill_date.strftime("%Y-%m-%d")
        else:
            date_str = "N/A"

        # Add Bill node
        if bill_title not in added_nodes:
            # Include the date in the tooltip
            tooltip = f"<b>Bill</b>: {bill_title}<br>Date: {date_str}"
            if link:
                tooltip += f"<br><a href='{link}' target='_blank'>Open Link</a>"

            net.add_node(bill_title, label=bill_title, title=tooltip, color="#ffa500")
            added_nodes.add(bill_title)

        # Add Author node
        if author and author not in added_nodes:
            net.add_node(author, label=author, title=f"<b>Author</b>: {author}", color="#1f78b4")
            added_nodes.add(author)

        # Add Policy node
        if policy_area and policy_area not in added_nodes:
            net.add_node(policy_area, label=policy_area,
                         title=f"<b>Policy Area</b>: {policy_area}", color="#33a02c")
            added_nodes.add(policy_area)

        # Add edges
        if author:
            net.add_edge(author, bill_title, color="#bbbbbb")
        if policy_area:
            net.add_edge(bill_title, policy_area, color="#bbbbbb")

    return net

def render_network_graph_with_dblclick(net: Network, data: pd.DataFrame):
    """
    Renders the PyVis network graph in Streamlit by generating
    an HTML file and embedding it via an iframe.

    Injects custom JS so double-clicking a Bill node opens its link in a new tab.
    """
    # 1) Build node->link map for Bill nodes only
    node_link_map = {}
    for idx, row in data.iterrows():
        bill_title = row["Title"]
        link = row.get("Link", None)
        if pd.notna(link) and link:
            node_link_map[bill_title] = link

    # 2) Generate the HTML in memory
    net.generate_html()

    # 3) Inject custom <script> to handle double-click
    custom_js = f"""
<script>
var linkMap = {json.dumps(node_link_map)};
function openLinkOnDoubleClick(params) {{
    if (params.nodes.length > 0) {{
        var nodeId = params.nodes[0];
        if (linkMap[nodeId]) {{
            window.open(linkMap[nodeId], '_blank');
        }}
    }}
}}
document.addEventListener('DOMContentLoaded', function() {{
    if (typeof network !== 'undefined') {{
        network.on("doubleClick", openLinkOnDoubleClick);
    }}
}});
</script>
"""
    if net.html is not None:
        net.html = net.html.replace("</body>", custom_js + "</body>")

    # 4) Write final HTML to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".html") as tmp_file:
        temp_path = tmp_file.name

    with open(temp_path, "w", encoding="utf-8") as f:
        f.write(net.html)

    # 5) Display in an iframe
    with open(temp_path, "r", encoding="utf-8") as f:
        html_content = f.read()
    components.html(html_content, height=700, scrolling=True)

def create_sankey_diagram(df: pd.DataFrame):
    """
    Creates a Sankey diagram with improved styling to reduce clutter.
    It visualizes the flow from Author -> Policy Area -> Enactment Method
    using plotly.graph_objects.
    """
    authors = sorted(df["Author"].dropna().unique())
    policies = sorted(df["Policy Area"].dropna().unique())
    methods = sorted(df["Enactment Method"].dropna().unique())

    labels = list(authors) + list(policies) + list(methods)

    author_indices = {a: i for i, a in enumerate(authors)}
    policy_indices = {p: i + len(authors) for i, p in enumerate(policies)}
    method_indices = {
        m: i + len(authors) + len(policies)
        for i, m in enumerate(methods)
    }

    sources = []
    targets = []
    values = []

    for _, row in df.iterrows():
        author = row["Author"]
        policy = row["Policy Area"]
        if pd.notna(author) and pd.notna(policy):
            sources.append(author_indices[author])
            targets.append(policy_indices[policy])
            values.append(1)

    for _, row in df.iterrows():
        policy = row["Policy Area"]
        method = row["Enactment Method"]
        if pd.notna(policy) and pd.notna(method):
            sources.append(policy_indices[policy])
            targets.append(method_indices[method])
            values.append(1)

    fig = go.Figure(
        data=[
            go.Sankey(
                arrangement="snap",
                node=dict(
                    pad=20,
                    thickness=20,
                    line=dict(color="#333", width=0.5),
                    label=labels,
                    color="#666",
                    hovertemplate='%{label}<extra></extra>',
                ),
                link=dict(
                    source=sources,
                    target=targets,
                    value=values,
                    color="rgba(150,150,150,0.4)",
                    hovertemplate=(
                        'Flow from %{source.label} to %{target.label} '
                        'has a value of %{value}<extra></extra>'
                    ),
                ),
            )
        ]
    )

    fig.update_layout(
        title_text="Sankey: Author → Policy Area → Method",
        font=dict(size=14),
        height=600,
        margin=dict(l=50, r=50, t=50, b=50),
    )

    return fig

def create_timeline_plot(df: pd.DataFrame):
    """
    Creates a Plotly timeline showing the bills over time.
    We replicate a timeline by setting Start = Date and End = Date+1 day.
    """
    temp_df = df.copy()
    temp_df["Start"] = temp_df["Date"]
    temp_df["End"] = temp_df["Date"] + pd.Timedelta(days=1)

    fig = px.timeline(
        temp_df,
        x_start="Start",
        x_end="End",
        y="Title",
        color="Author",
        hover_data=["Policy Area", "Enactment Method", "Link"],
        title="Timeline of Bills (1-Day Window)"
    )
    fig.update_yaxes(autorange="reversed")  # earliest item at top
    fig.update_layout(height=700)
    return fig

# ==========================================
#       MAIN APP
# ==========================================
def main():
    st.set_page_config(page_title="Enacted Federal Legislation Tracker", layout="wide")

    # ----- Sidebar with info -----
    with st.sidebar:
        st.title("About this App")
        st.markdown("""
        **Enacted Federal Legislation Tracker**  
        Version 2.0.

        This enhanced app includes:
        - **Filters** by Author, Policy, Enactment Method, and Date.
        - **Network Graph** with physics-based layout (PyVis).  
          (Double-click Bill nodes to open links!)
        - **Scatter & Bar Charts** with advanced mode.
        - **Sankey Diagram** for flow-based analysis.
        - **Timeline** visualization.

        *Tip*: Adjust filters to narrow down data before using heavier visuals!
        """)
        st.info(
            "Make sure your data file is in the same folder.\n\n"
            "Hover over Bill nodes to see their date.\n"
            "Double-click a Bill node to open its link in a new tab!"
        )

    st.title("Enacted Federal Legislation Tracker")

    # 1. Load Data
    data = get_filtered_data()
    if data.empty:
        st.error("No data available. Please ensure the file is present and correctly formatted.")
        return

    # 2. "See All" Button for entire dataset
    show_all = st.button("See all bills (Warning: might take a minute to load)")
    if show_all:
        filtered_data = data
        st.warning("Showing ALL bills. This may be slow if the dataset is large.")
    else:
        # -- By default, only show SULLIVAN authorship to start --
        st.subheader("Search / Filter Options")

        search_option = st.radio(
            "How would you like to search for bills?",
            ["Author", "Method of Enactment", "Policy Area", "Date Range"],
            index=0
        )

        authors = sorted(data["Author"].dropna().unique())
        policy_areas = sorted(data["Policy Area"].dropna().unique())
        methods = sorted(data["Enactment Method"].dropna().unique())
        min_date = data["Date"].min().date()
        max_date = data["Date"].max().date()

        # If default author exists in dataset, pre-select it
        default_author = [DEFAULT_AUTHOR_NAME] if DEFAULT_AUTHOR_NAME in authors else []

        author_filter = []
        policy_filter = []
        enactment_filter = []
        date_range = (min_date, max_date)

        if search_option == "Author":
            author_filter = st.multiselect(
                "Select Author(s)",
                options=authors,
                default=default_author
            )
        elif search_option == "Method of Enactment":
            enactment_filter = st.multiselect("Select Method(s) of Enactment", options=methods, default=[])
        elif search_option == "Policy Area":
            policy_filter = st.multiselect("Select Policy Area(s)", options=policy_areas, default=[])
        elif search_option == "Date Range":
            date_range = st.slider(
                "Select Date Range",
                min_value=min_date,
                max_value=max_date,
                value=(min_date, max_date)
            )

        filtered_data = data[
            ((data["Author"].isin(author_filter)) | (len(author_filter) == 0)) &
            ((data["Policy Area"].isin(policy_filter)) | (len(policy_filter) == 0)) &
            ((data["Enactment Method"].isin(enactment_filter)) | (len(enactment_filter) == 0)) &
            (data["Date"] >= pd.to_datetime(date_range[0])) &
            (data["Date"] <= pd.to_datetime(date_range[1]))
        ]

    # -- PHYSICS GRAPH FIRST --
    st.subheader("Network Graph (Physics Simulation)")
    st.markdown(
        """
        **Hover over Bill nodes** to see their date.  
        **Double-click on Bill nodes** to open the link in a new tab!  
        
        - **Author** nodes (blue)  
        - **Bill** nodes (orange)  
        - **Policy Area** nodes (green)

        Drag them around to see the “springs” in action.
        """
    )
    if not filtered_data.empty:
        net = create_network_graph(filtered_data)
        render_network_graph_with_dblclick(net, filtered_data)
    else:
        st.info("No data to display in network graph. Adjust your filters or load all bills.")

    # -- FILTERED RESULTS TABLE --
    st.subheader("Filtered Results")
    display_results_table(filtered_data)

    # -- BASIC SCATTER PLOT --
    if not filtered_data.empty:
        st.subheader("Basic Visualization")
        st.markdown(
            """
            **Tip**: Use the Plotly toolbar (top-right) to zoom, pan, or go full screen.
            **Click** on a legend entry to hide/show certain series.
            """
        )
        fig = generate_scatter_plot(
            data=filtered_data,
            x_col="Policy Area",
            y_col="Date",
            color_col="Author",
            title="Policy Area vs. Date (Colored by Author)"
        )
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("No data to visualize. Please adjust filters or load all bills.")

    # -- BAR CHART BY YEAR (Expander) --
    with st.expander("Show Bar Chart by Year", expanded=False):
        if not filtered_data.empty:
            temp_df = filtered_data.copy()
            temp_df["Year"] = temp_df["Date"].dt.year
            year_counts = temp_df.groupby("Year")["Title"].count().reset_index()
            if not year_counts.empty:
                st.write("Number of Enacted Items per Year")
                fig_bar = px.bar(
                    year_counts,
                    x="Year",
                    y="Title",
                    labels={"Title": "Count of Enacted Items"},
                    title="Enacted Items by Year",
                )
                st.plotly_chart(fig_bar, use_container_width=True)
            else:
                st.info("No data for bar chart. Adjust your filters.")
        else:
            st.info("No data available. Please adjust filters or load all bills.")

    # -- ADVANCED MODE --
    with st.expander("Show Advanced Filters and Visualization", expanded=False):
        st.markdown(
            """
            **Advanced Mode**:
            - Combine multiple filters at once (e.g., multiple authors **AND** multiple policy areas).
            - Choose columns for X-axis, Y-axis, and color dimension.
            - Adjust font sizes and toggle clickable labels for each orb.
            """
        )
        if show_all:
            st.info("You are currently viewing ALL bills. Advanced filtering won't reduce data.")
            adv_data = filtered_data
        else:
            # Let user choose advanced filters if not showing all
            authors = sorted(data["Author"].dropna().unique())
            policy_areas = sorted(data["Policy Area"].dropna().unique())
            methods = sorted(data["Enactment Method"].dropna().unique())
            min_date = data["Date"].min().date()
            max_date = data["Date"].max().date()

            advanced_author_filter = st.multiselect("Filter by Author", options=authors, default=[])
            advanced_policy_filter = st.multiselect("Filter by Policy Area", options=policy_areas, default=[])
            advanced_enactment_filter = st.multiselect("Filter by Enactment Method", options=methods, default=[])
            advanced_date_range = st.slider(
                "Select Date Range",
                min_value=min_date,
                max_value=max_date,
                value=(min_date, max_date)
            )

            adv_data = data[
                ((data["Author"].isin(advanced_author_filter)) | (len(advanced_author_filter) == 0)) &
                ((data["Policy Area"].isin(advanced_policy_filter)) | (len(advanced_policy_filter) == 0)) &
                ((data["Enactment Method"].isin(advanced_enactment_filter)) | (len(advanced_enactment_filter) == 0)) &
                (data["Date"] >= pd.to_datetime(advanced_date_range[0])) &
                (data["Date"] <= pd.to_datetime(advanced_date_range[1]))
            ]

        st.subheader("Advanced Filtered Results")
        display_results_table(adv_data)

        axis_options = ["Policy Area", "Date", "Author", "Enactment Method"]
        x_axis = st.selectbox("X-Axis", axis_options, index=0)
        y_axis = st.selectbox("Y-Axis", axis_options, index=1)
        color_col = st.selectbox("Color By", axis_options, index=2)

        text_size = st.slider("Text Size in Chart", min_value=10, max_value=30, value=12, step=1)
        annotate_advanced = st.checkbox("Show Titles (Clickable) Above Each Orb?", value=True)

        if not adv_data.empty:
            fig_advanced = generate_scatter_plot(
                data=adv_data,
                x_col=x_axis,
                y_col=y_axis,
                color_col=color_col,
                title="Advanced Visualization",
                text_size=text_size,
                annotate_points=annotate_advanced
            )
            st.plotly_chart(fig_advanced, use_container_width=True)
        else:
            st.info("No data to visualize in Advanced Mode. Please adjust filters or load all bills.")

    # -- SANKEY DIAGRAM --
    with st.expander("Sankey Diagram (Author → Policy Area → Method)"):
        if not filtered_data.empty:
            fig_sankey = create_sankey_diagram(filtered_data)
            st.plotly_chart(fig_sankey, use_container_width=True)
        else:
            st.info("No data to display for Sankey. Adjust your filters or load all bills.")

    # -- TIMELINE VIEW --
    with st.expander("Timeline View"):
        if not filtered_data.empty:
            fig_timeline = create_timeline_plot(filtered_data)
            st.plotly_chart(fig_timeline, use_container_width=True)
        else:
            st.info("No data to display in Timeline. Please adjust filters or load all bills.")

# ==========================================
#     RUN THE APP
# ==========================================
if __name__ == "__main__":
    main()

